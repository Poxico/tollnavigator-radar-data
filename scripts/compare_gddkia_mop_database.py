#!/usr/bin/env python3
"""Build and compare a conservative TollNavigator MOP candidate database from GDDKiA XLSX.

Diagnostic mode only: the script never overwrites the published production database.
It preserves app-only enrichment (brands/fuelBrands), never deletes an old record merely
because it is absent from the workbook, and reports all potentially destructive changes.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

PAGE_URL = "https://www.gov.pl/web/gddkia/wykaz-parkingow-i-mop"
REQUIRED_HEADERS = {
    "Dane identyfikacyjne",
    "Obecna Kategoria",
    "x",
    "y",
    "Klasa techniczna drogi",
    "Nr drogi",
    "Pikietaż",
    "Kierunek",
    "Parking Osobowe",
    "Parking Ciężarowe",
    "Parking Autobus",
    "Toalety",
    "Stacja paliw",
    "Restauracja /Bistro",
    "Miejsca noclegowe",
    "Ładowanie pojazdów elektrycznych",
    "Prysznic",
}
BOOLEAN_TO_APP = {
    "Toalety": "hasToilet",
    "Stacja paliw": "hasFuel",
    "Restauracja /Bistro": "hasRestaurant",
    "Miejsca noclegowe": "hasHotel",
    "Ładowanie pojazdów elektrycznych": "hasEV",
    "Prysznic": "hasShower",
}
CORE_FIELDS = (
    "name", "motorway", "km", "lat", "lon", "direction", "category",
    "parkingCar", "parkingTruck",
)
SERVICE_FIELDS = tuple(BOOLEAN_TO_APP.values())


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def folded(value: Any) -> str:
    result = unicodedata.normalize("NFKD", text(value)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", result.casefold()).strip()


def yes(value: Any) -> bool:
    return folded(value) in {"tak", "yes", "true", "1", "x"}


def number(value: Any) -> int:
    if value is None or text(value) == "":
        return 0
    try:
        return int(float(str(value).replace(",", ".")))
    except ValueError:
        match = re.search(r"-?\d+", text(value))
        return int(match.group()) if match else 0


def coordinate(value: Any) -> float:
    return round(float(str(value).replace(",", ".")), 7)


def kilometre(value: Any) -> float:
    raw = text(value).replace(" ", "").replace(",", ".")
    match = re.fullmatch(r"(\d+)[+](\d{1,3})", raw)
    if match:
        return round(int(match.group(1)) + int(match.group(2)) / 1000.0, 3)
    return round(float(raw), 3)


def road(road_class: Any, road_number: Any) -> str:
    klass = re.sub(r"[^A-Za-z]", "", text(road_class)).upper()
    number_text = re.sub(r"\s+", "", text(road_number)).upper().replace(".0", "")
    return f"{klass}{number_text}"


def safe_id(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    normalized = re.sub(r"[^A-Za-z0-9]+", "_", normalized.upper()).strip("_")
    return normalized


def generated_id(item: dict[str, Any]) -> str:
    km = str(item["km"]).replace(".", "_")
    return safe_id(f'{item["motorway"]}_{item["name"]}_{km}_{item["direction"]}')


def distance_m(a: dict[str, Any], b: dict[str, Any]) -> float:
    lat1, lon1 = math.radians(float(a["lat"])), math.radians(float(a["lon"]))
    lat2, lon2 = math.radians(float(b["lat"])), math.radians(float(b["lon"]))
    dlat, dlon = lat2 - lat1, lon2 - lon1
    h = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 6371000.0 * 2 * math.atan2(math.sqrt(h), math.sqrt(max(0.0, 1 - h)))


def exact_key(item: dict[str, Any]) -> tuple[str, str, int, str]:
    return (
        folded(item.get("motorway")),
        folded(item.get("name")),
        round(float(item.get("km", 0)) * 1000),
        folded(item.get("direction")),
    )


def route_key(item: dict[str, Any]) -> tuple[str, int, str]:
    return (
        folded(item.get("motorway")),
        round(float(item.get("km", 0)) * 1000),
        folded(item.get("direction")),
    )


def find_attachment(page_url: str) -> tuple[str, str, bytes]:
    response = requests.get(page_url, timeout=45, headers={"User-Agent": "TollNavigator-GDDKiA-MOP-updater/1.0"})
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    candidates: list[tuple[int, str, str]] = []
    for link in soup.find_all("a", href=True):
        href = str(link["href"])
        label = " ".join(link.stripped_strings)
        haystack = folded(f"{label} {href}")
        if "mop" not in haystack:
            continue
        score = 0
        if "xlsx" in haystack or "xls" in haystack:
            score += 20
        if "tabela" in haystack:
            score += 10
        if "marzec 2026" in haystack:
            score += 5
        if "/attachment/" in href:
            score += 3
        candidates.append((score, label, requests.compat.urljoin(page_url, href)))
    if not candidates:
        raise RuntimeError("No GDDKiA MOP workbook attachment found")
    _, label, url = max(candidates, key=lambda item: item[0])
    workbook = requests.get(url, timeout=60, headers={"User-Agent": "TollNavigator-GDDKiA-MOP-updater/1.0"})
    workbook.raise_for_status()
    if not workbook.content.startswith(b"PK"):
        raise RuntimeError(f"Downloaded attachment is not an XLSX file: {workbook.headers.get('content-type')}")
    return label, url, workbook.content


def parse_workbook(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if len(wb.sheetnames) != 1:
        raise RuntimeError(f"Expected one worksheet, found {len(wb.sheetnames)}")
    ws = wb[wb.sheetnames[0]]
    headers = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = sorted(REQUIRED_HEADERS - set(headers))
    if missing:
        raise RuntimeError(f"Missing required GDDKiA columns: {missing}")
    index = {header: position for position, header in enumerate(headers)}
    items: list[dict[str, Any]] = []
    row_issues: list[dict[str, Any]] = []
    target_categories: Counter[str] = Counter()
    for row_number, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(cells)
        if not any(value is not None and text(value) for value in values):
            continue
        try:
            item = {
                "sourceRow": row_number,
                "name": text(values[index["Dane identyfikacyjne"]]) or text(values[index.get("Miejscowość", 0)]),
                "motorway": road(values[index["Klasa techniczna drogi"]], values[index["Nr drogi"]]),
                "km": kilometre(values[index["Pikietaż"]]),
                "lat": coordinate(values[index["x"]]),
                "lon": coordinate(values[index["y"]]),
                "direction": text(values[index["Kierunek"]]),
                "category": text(values[index["Obecna Kategoria"]]),
                "parkingCar": number(values[index["Parking Osobowe"]]),
                "parkingTruck": number(values[index["Parking Ciężarowe"]]),
                "parkingBus": number(values[index["Parking Autobus"]]),
                "targetCategory": text(values[index["Kat_docel"]]) if "Kat_docel" in index else "",
                "sourceManager": text(values[index["Zarządca"]]) if "Zarządca" in index else "",
                "sourceTown": text(values[index["Miejscowość"]]) if "Miejscowość" in index else "",
            }
            for source, app in BOOLEAN_TO_APP.items():
                item[app] = yes(values[index[source]])
            if not item["name"] or not item["motorway"] or not item["direction"]:
                raise ValueError("missing name, road or direction")
            target_categories[item["targetCategory"]] += 1
            items.append(item)
        except Exception as exc:
            row_issues.append({"row": row_number, "error": f"{type(exc).__name__}: {exc}"})
    diagnostics = {
        "sheet": ws.title,
        "rowCount": ws.max_row,
        "columnCount": ws.max_column,
        "headers": headers,
        "parsedCount": len(items),
        "rowIssues": row_issues,
        "targetCategories": dict(sorted(target_categories.items())),
    }
    return items, diagnostics


def match_items(source_items: list[dict[str, Any]], previous: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    exact_map: dict[tuple[str, str, int, str], list[int]] = defaultdict(list)
    route_map: dict[tuple[str, int, str], list[int]] = defaultdict(list)
    road_map: dict[str, list[int]] = defaultdict(list)
    for idx, old in enumerate(previous):
        exact_map[exact_key(old)].append(idx)
        route_map[route_key(old)].append(idx)
        road_map[folded(old.get("motorway"))].append(idx)

    used: set[int] = set()
    candidates: list[dict[str, Any]] = []
    match_counts: Counter[str] = Counter()
    unmatched_source: list[dict[str, Any]] = []
    changes: list[dict[str, Any]] = []
    destructive: list[dict[str, Any]] = []

    for source in source_items:
        matched_idx: int | None = None
        method = "new"
        choices = [idx for idx in exact_map.get(exact_key(source), []) if idx not in used]
        if len(choices) == 1:
            matched_idx, method = choices[0], "exact"
        if matched_idx is None:
            choices = [idx for idx in route_map.get(route_key(source), []) if idx not in used]
            if len(choices) == 1:
                matched_idx, method = choices[0], "road_km_direction"
        if matched_idx is None:
            nearby: list[tuple[float, int]] = []
            for idx in road_map.get(folded(source["motorway"]), []):
                if idx in used:
                    continue
                dist = distance_m(source, previous[idx])
                if dist <= 350:
                    nearby.append((dist, idx))
            nearby.sort()
            if nearby and (len(nearby) == 1 or nearby[1][0] - nearby[0][0] >= 75):
                matched_idx, method = nearby[0][1], "coordinate"

        if matched_idx is None:
            new_item = {key: source[key] for key in CORE_FIELDS + SERVICE_FIELDS}
            new_item["id"] = generated_id(new_item)
            new_item["brands"] = []
            new_item["fuelBrands"] = []
            candidates.append(new_item)
            unmatched_source.append({key: source.get(key) for key in ("sourceRow", "name", "motorway", "km", "direction", "lat", "lon")})
            match_counts["new"] += 1
            continue

        used.add(matched_idx)
        old = previous[matched_idx]
        candidate = dict(old)
        record_changes: dict[str, Any] = {}
        record_destructive: dict[str, Any] = {}
        for field in CORE_FIELDS:
            new_value = source[field]
            old_value = old.get(field)
            if new_value != old_value:
                record_changes[field] = {"old": old_value, "new": new_value}
                candidate[field] = new_value
        for field in SERVICE_FIELDS:
            new_value = bool(source[field])
            old_value = bool(old.get(field))
            if new_value and not old_value:
                record_changes[field] = {"old": old_value, "new": new_value}
                candidate[field] = True
            elif old_value and not new_value:
                record_destructive[field] = {"old": True, "source": False, "applied": False}
                candidate[field] = True
        candidate["id"] = old.get("id") or generated_id(candidate)
        candidate["brands"] = list(old.get("brands") or [])
        candidate["fuelBrands"] = list(old.get("fuelBrands") or [])
        candidates.append(candidate)
        match_counts[method] += 1
        if record_changes:
            changes.append({
                "id": candidate["id"], "name": candidate["name"], "method": method,
                "sourceRow": source["sourceRow"], "changes": record_changes,
                "parkingBus": source["parkingBus"], "targetCategory": source["targetCategory"],
            })
        if record_destructive:
            destructive.append({
                "id": candidate["id"], "name": candidate["name"], "method": method,
                "sourceRow": source["sourceRow"], "blockedChanges": record_destructive,
            })

    preserved_missing: list[dict[str, Any]] = []
    for idx, old in enumerate(previous):
        if idx not in used:
            candidates.append(dict(old))
            preserved_missing.append({
                "id": old.get("id"), "name": old.get("name"), "motorway": old.get("motorway"),
                "km": old.get("km"), "direction": old.get("direction"),
            })

    candidate_ids = [item.get("id") for item in candidates]
    duplicate_ids = sorted(key for key, count in Counter(candidate_ids).items() if count > 1)
    diagnostics = {
        "matchCounts": dict(sorted(match_counts.items())),
        "sourceCount": len(source_items),
        "previousCount": len(previous),
        "candidateCount": len(candidates),
        "matchedPreviousCount": len(used),
        "newSourceCount": len(unmatched_source),
        "preservedMissingPreviousCount": len(preserved_missing),
        "changedRecordCount": len(changes),
        "blockedDestructiveRecordCount": len(destructive),
        "duplicateIds": duplicate_ids,
        "unmatchedSource": unmatched_source,
        "preservedMissingPrevious": preserved_missing,
        "changes": changes,
        "blockedDestructiveChanges": destructive,
    }
    return candidates, diagnostics


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--previous", required=True)
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--candidate", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--summary", required=True)
    parser.add_argument("--page-url", default=PAGE_URL)
    args = parser.parse_args()

    previous_root = json.loads(Path(args.previous).read_text(encoding="utf-8"))
    previous = previous_root.get("rest_areas") or []
    label, attachment_url, workbook_bytes = find_attachment(args.page_url)
    xlsx_path = Path(args.xlsx)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path.write_bytes(workbook_bytes)
    source_items, workbook_diagnostics = parse_workbook(xlsx_path)
    candidates, comparison = match_items(source_items, previous)

    candidate_root = {
        "version": datetime.now(timezone.utc).date().isoformat(),
        "generated": utc_now(),
        "source": "GDDKiA official MOP workbook — diagnostic candidate",
        "count": len(candidates),
        "rest_areas": candidates,
    }
    candidate_path = Path(args.candidate)
    candidate_path.parent.mkdir(parents=True, exist_ok=True)
    candidate_path.write_text(json.dumps(candidate_root, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    report = {
        "schemaVersion": 1,
        "generatedAt": utc_now(),
        "mode": "diagnostic_only",
        "productionDatabaseModified": False,
        "source": {
            "pageUrl": args.page_url,
            "attachmentUrl": attachment_url,
            "attachmentLabel": label,
            "size": len(workbook_bytes),
            "sha256": hashlib.sha256(workbook_bytes).hexdigest(),
        },
        "workbook": workbook_diagnostics,
        "comparison": comparison,
        "safeToConsiderPublishing": (
            workbook_diagnostics["parsedCount"] >= 350
            and not workbook_diagnostics["rowIssues"]
            and not comparison["duplicateIds"]
            and comparison["matchedPreviousCount"] >= 350
            and comparison["candidateCount"] >= comparison["previousCount"] * 0.95
        ),
    }
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    summary = [
        "## TollNavigator — porównanie bazy MOP z GDDKiA",
        "",
        "| Kontrola | Wynik |",
        "|---|---:|",
        f"| Wiersze danych GDDKiA | {workbook_diagnostics['parsedCount']} |",
        f"| Poprzednia dobra baza | {comparison['previousCount']} |",
        f"| Dopasowane rekordy | {comparison['matchedPreviousCount']} |",
        f"| Nowe rekordy GDDKiA | {comparison['newSourceCount']} |",
        f"| Stare zachowane mimo braku w arkuszu | {comparison['preservedMissingPreviousCount']} |",
        f"| Rekordy z bezpiecznie zastosowanymi zmianami | {comparison['changedRecordCount']} |",
        f"| Rekordy z zablokowanym pogorszeniem usług | {comparison['blockedDestructiveRecordCount']} |",
        f"| Kandydacka baza | {comparison['candidateCount']} |",
        "",
        f"**Ocena diagnostyczna:** `{'PASS' if report['safeToConsiderPublishing'] else 'REVIEW'}`",
        "",
        "> Produkcyjna baza nie została zmieniona. Brak wiersza w Excelu nie usuwa MOP-u, marki są zachowywane, a zmiany `tak → nie` są tylko raportowane.",
    ]
    Path(args.summary).write_text("\n".join(summary) + "\n", encoding="utf-8")
    print("\n".join(summary))


if __name__ == "__main__":
    main()
