#!/usr/bin/env python3
"""Download and inspect the current official GDDKiA MOP workbook.

This is a diagnostic-only step. It never modifies the production JSON database.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

SOURCE_PAGE = "https://www.gov.pl/web/gddkia/wykaz-parkingow-i-mop"
USER_AGENT = "TollNavigator-GDDKiA-MOP-updater/1.0"

KEYWORDS = {
    "road": ("droga", "nr drogi", "numer drogi", "trasa"),
    "name": ("nazwa", "mop", "parking", "obiekt"),
    "chainage": ("pikietaz", "kilometraz", "kilometr", "km"),
    "direction": ("kierunek", "strona", "jezdnia"),
    "category": ("kategoria", "funkcja", "typ mop", "rodzaj"),
    "cars": ("osobow", "samochody osobowe", "miejsca osobowe"),
    "trucks": ("ciezar", "hgv", "tir", "samochody ciezarowe"),
    "buses": ("autobus", "autokar"),
    "fuel": ("paliw", "stacja"),
    "food": ("gastronom", "restaur", "bar"),
    "hotel": ("hotel", "motel", "nocleg"),
    "coordinates": ("wspolrzed", "szerokosc", "dlugosc", "latitude", "longitude"),
}


def now_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def normalize(value: Any) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower().replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", text).strip()


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def download_source(session: requests.Session, page_url: str, target: Path) -> dict[str, Any]:
    page = session.get(page_url, timeout=45)
    page.raise_for_status()
    soup = BeautifulSoup(page.text, "html.parser")

    candidates: list[tuple[int, str, str]] = []
    for anchor in soup.find_all("a", href=True):
        href = urljoin(page_url, anchor.get("href", ""))
        label = " ".join(anchor.stripped_strings)
        normalized = normalize(label + " " + href)
        score = 0
        if "tabela mop" in normalized:
            score += 100
        if "istniej" in normalized:
            score += 30
        if ".xlsx" in normalized:
            score += 20
        if "/attachment/" in href:
            score += 10
        if score:
            candidates.append((score, href, label))

    if not candidates:
        raise RuntimeError("No GDDKiA MOP XLSX attachment link found on the official page")

    candidates.sort(key=lambda item: item[0], reverse=True)
    _, workbook_url, label = candidates[0]
    response = session.get(workbook_url, timeout=60)
    response.raise_for_status()
    if len(response.content) < 10_000:
        raise RuntimeError(f"Downloaded workbook is suspiciously small: {len(response.content)} bytes")
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(response.content)

    return {
        "pageUrl": page_url,
        "attachmentUrl": workbook_url,
        "attachmentLabel": label,
        "contentType": response.headers.get("content-type"),
        "contentLength": len(response.content),
        "sha256": hashlib.sha256(response.content).hexdigest(),
        "etag": response.headers.get("etag"),
        "lastModified": response.headers.get("last-modified"),
    }


def row_signature(values: list[Any]) -> dict[str, Any]:
    normalized_cells = [normalize(value) for value in values]
    joined = " | ".join(cell for cell in normalized_cells if cell)
    matches: dict[str, list[str]] = {}
    for category, needles in KEYWORDS.items():
        found = [needle for needle in needles if normalize(needle) in joined]
        if found:
            matches[category] = found
    core = sum(1 for key in ("road", "name", "chainage", "direction", "category") if key in matches)
    parking = sum(1 for key in ("cars", "trucks", "buses") if key in matches)
    score = core * 10 + parking * 6 + len(matches)
    return {"score": score, "matches": matches, "text": joined[:1500]}


def inspect_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheets: list[dict[str, Any]] = []

    for sheet in workbook.worksheets:
        rows: list[dict[str, Any]] = []
        header_candidates: list[dict[str, Any]] = []
        max_scan_row = min(sheet.max_row, 60)
        max_scan_col = min(sheet.max_column, 80)

        for row_index in range(1, max_scan_row + 1):
            values = [json_value(sheet.cell(row_index, col).value) for col in range(1, max_scan_col + 1)]
            nonempty = [value for value in values if value not in (None, "")]
            if nonempty:
                rows.append({"row": row_index, "values": values})
                signature = row_signature(values)
                if signature["score"] > 0:
                    header_candidates.append({"row": row_index, **signature, "values": values})

        header_candidates.sort(key=lambda item: (item["score"], -item["row"]), reverse=True)
        sheets.append(
            {
                "title": sheet.title,
                "maxRow": sheet.max_row,
                "maxColumn": sheet.max_column,
                "mergedRanges": [str(item) for item in list(sheet.merged_cells.ranges)[:100]],
                "freezePanes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
                "autoFilter": str(sheet.auto_filter.ref) if sheet.auto_filter and sheet.auto_filter.ref else None,
                "sampleRows": rows[:35],
                "headerCandidates": header_candidates[:10],
            }
        )

    return {"sheetCount": len(sheets), "sheetNames": workbook.sheetnames, "sheets": sheets}


def write_summary(report: dict[str, Any], path: Path | None) -> None:
    lines = [
        "## TollNavigator — analiza arkusza GDDKiA MOP",
        "",
        f"- Załącznik: `{report['source']['attachmentLabel']}`",
        f"- Rozmiar: **{report['source']['contentLength']} B**",
        f"- SHA-256: `{report['source']['sha256']}`",
        f"- Arkusze: **{report['workbook']['sheetCount']}** — {', '.join(report['workbook']['sheetNames'])}",
        "",
        "### Najlepsi kandydaci na wiersz nagłówków",
        "",
        "| Arkusz | Wiersz | Wynik | Rozpoznane pola |",
        "|---|---:|---:|---|",
    ]
    for sheet in report["workbook"]["sheets"]:
        for candidate in sheet["headerCandidates"][:3]:
            fields = ", ".join(candidate["matches"].keys())
            lines.append(f"| {sheet['title']} | {candidate['row']} | {candidate['score']} | {fields} |")
    text = "\n".join(lines) + "\n"
    if path:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(text, encoding="utf-8")
    print(text)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--page", default=SOURCE_PAGE)
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--summary")
    args = parser.parse_args()

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT, "Accept-Language": "pl-PL,pl;q=0.9,en;q=0.5"})

    xlsx_path = Path(args.xlsx)
    source = download_source(session, args.page, xlsx_path)
    workbook = inspect_workbook(xlsx_path)
    report = {
        "schemaVersion": 1,
        "generatedAt": now_utc(),
        "mode": "diagnostic_only",
        "productionDatabaseModified": False,
        "source": source,
        "workbook": workbook,
    }
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_summary(report, Path(args.summary) if args.summary else None)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        raise
